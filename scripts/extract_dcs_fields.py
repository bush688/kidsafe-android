from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Font


def _norm(s: object) -> str:
    if s is None:
        return ""
    if isinstance(s, str):
        return re.sub(r"\s+", "", s)
    return re.sub(r"\s+", "", str(s))


def _norm_keep_spaces(s: object) -> str:
    if s is None:
        return ""
    if isinstance(s, str):
        return re.sub(r"[ \t\r\n]+", " ", s).strip()
    return re.sub(r"[ \t\r\n]+", " ", str(s)).strip()


@dataclass(frozen=True)
class Record:
    instrument_tag: str
    purpose: str
    measure_range: str
    unit: str


TARGET_HEADERS = {
    "仪表位号": ("instrument_tag",),
    "用途": ("purpose",),
    "测量范围": ("measure_range",),
    "工程单位": ("unit",),
}


def _find_existing_path(p: Path) -> Optional[Path]:
    if p.exists():
        return p
    if p.suffix.lower() == ".pdf":
        parent = p.parent
        if parent.exists():
            name_norm = _norm(p.name)
            for candidate in parent.glob("*.pdf"):
                if _norm(candidate.name) == name_norm:
                    return candidate
    return None


def _guess_pdf_in_folder(folder: Path) -> Optional[Path]:
    if not folder.exists():
        return None
    pdfs = sorted(folder.glob("*.pdf"))
    for p in pdfs:
        if "控制系统监控数据表" in p.name and "DCS" in p.name:
            return p
    return pdfs[0] if pdfs else None


def _extract_from_pdf(pdf_path: Path) -> list[Record]:
    wanted_norm = {k: _norm(k) for k in TARGET_HEADERS.keys()}
    records: list[Record] = []
    seen_header = False
    col_map: dict[str, int] = {}

    def consider_row(row: list[object]) -> None:
        nonlocal seen_header, col_map, records
        if not row:
            return
        cells = [_norm_keep_spaces(c) for c in row]
        cells_norm = [_norm(c) for c in row]

        if not seen_header:
            for i, c in enumerate(cells_norm):
                for header, header_norm in wanted_norm.items():
                    if c == header_norm and header not in col_map:
                        col_map[header] = i
            if len(col_map) == 4:
                seen_header = True
            return

        if not any(cells_norm):
            return

        def get(header: str) -> str:
            idx = col_map.get(header)
            if idx is None or idx >= len(cells):
                return ""
            return cells[idx].strip()

        instrument_tag = get("仪表位号")
        if not instrument_tag:
            return

        records.append(
            Record(
                instrument_tag=instrument_tag,
                purpose=get("用途"),
                measure_range=get("测量范围"),
                unit=get("工程单位"),
            )
        )

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables(
                {
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "intersection_tolerance": 5,
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                    "edge_min_length": 10,
                    "min_words_vertical": 1,
                    "min_words_horizontal": 1,
                }
            )
            for t in tables or []:
                for row in t or []:
                    consider_row(row)

            if not seen_header:
                text = page.extract_text() or ""
                if all(h in text for h in TARGET_HEADERS.keys()):
                    pass

    dedup: dict[str, Record] = {}
    for r in records:
        k = _norm(r.instrument_tag)
        if k and k not in dedup:
            dedup[k] = r
    return list(dedup.values())


def _extract_from_xlsx(xlsx_path: Path) -> list[Record]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    header_row_idx: Optional[int] = None
    col_idx: dict[str, int] = {}
    wanted = list(TARGET_HEADERS.keys())

    max_scan_rows = min(ws.max_row, 80)
    max_scan_cols = min(ws.max_column, 80)

    for r in range(1, max_scan_rows + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_scan_cols + 1)]
        row_norm = [_norm(v) for v in row_vals]
        tmp: dict[str, int] = {}
        for i, cell in enumerate(row_norm, start=1):
            for header in wanted:
                if cell == _norm(header):
                    tmp[header] = i
        if len(tmp) == 4:
            header_row_idx = r
            col_idx = tmp
            break

    if header_row_idx is None:
        raise RuntimeError(f"未在 {xlsx_path} 中找到包含四个目标字段的表头行。")

    records: list[Record] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        instrument = _norm_keep_spaces(ws.cell(row=r, column=col_idx["仪表位号"]).value)
        if not instrument:
            continue
        records.append(
            Record(
                instrument_tag=instrument,
                purpose=_norm_keep_spaces(ws.cell(row=r, column=col_idx["用途"]).value),
                measure_range=_norm_keep_spaces(ws.cell(row=r, column=col_idx["测量范围"]).value),
                unit=_norm_keep_spaces(ws.cell(row=r, column=col_idx["工程单位"]).value),
            )
        )

    dedup: dict[str, Record] = {}
    for r in records:
        k = _norm(r.instrument_tag)
        if k and k not in dedup:
            dedup[k] = r
    return list(dedup.values())


def _write_excel(out_path: Path, records: Iterable[Record]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "提取结果"

    headers = ["仪表位号", "用途", "测量范围", "工程单位"]
    ws.append(headers)

    for r in records:
        ws.append([r.instrument_tag, r.purpose, r.measure_range, r.unit])

    header_font = Font(bold=True)
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    widths = [18, 24, 24, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", type=str, required=True)
    ap.add_argument("--xlsx_fallback", type=str, default="")
    ap.add_argument("--out", type=str, required=True)
    args = ap.parse_args()

    pdf_path = Path(args.pdf)
    pdf_path2 = _find_existing_path(pdf_path)
    if pdf_path2 is None:
        guessed = _guess_pdf_in_folder(pdf_path.parent)
        if guessed is None:
            raise FileNotFoundError(f"未找到PDF：{pdf_path}")
        pdf_path2 = guessed

    pdf_records = _extract_from_pdf(pdf_path2)
    pdf_map = {_norm(r.instrument_tag): r for r in pdf_records if _norm(r.instrument_tag)}

    xlsx_records: list[Record] = []
    xlsx_map: dict[str, Record] = {}
    xlsx_path2: Optional[Path] = None
    if args.xlsx_fallback:
        xlsx_path = Path(args.xlsx_fallback)
        xlsx_path2 = _find_existing_path(xlsx_path) or xlsx_path
        if xlsx_path2.exists():
            xlsx_records = _extract_from_xlsx(xlsx_path2)
            xlsx_map = {_norm(r.instrument_tag): r for r in xlsx_records if _norm(r.instrument_tag)}

    if not pdf_map and not xlsx_map:
        raise RuntimeError("未能从PDF/Excel中提取到记录（可能是扫描件或表格结构特殊）。")

    merged: dict[str, Record] = {}
    keys = set(pdf_map) | set(xlsx_map)
    for k in keys:
        pr = pdf_map.get(k)
        xr = xlsx_map.get(k)
        instrument_tag = (xr.instrument_tag if xr else pr.instrument_tag) if (xr or pr) else ""
        purpose = (xr.purpose if xr and xr.purpose else (pr.purpose if pr else ""))
        measure_range = (
            xr.measure_range
            if xr and xr.measure_range
            else (pr.measure_range if pr else "")
        )
        unit = (xr.unit if xr and xr.unit else (pr.unit if pr else ""))
        merged[k] = Record(
            instrument_tag=instrument_tag,
            purpose=purpose,
            measure_range=measure_range,
            unit=unit,
        )

    records_sorted = sorted(merged.values(), key=lambda r: _norm(r.instrument_tag))
    _write_excel(Path(args.out), records_sorted)

    overlap = set(pdf_map) & set(xlsx_map)
    mismatch = 0
    for k in overlap:
        a = pdf_map[k]
        b = xlsx_map[k]
        if (a.purpose and b.purpose and a.purpose != b.purpose) or (
            a.measure_range and b.measure_range and a.measure_range != b.measure_range
        ) or (a.unit and b.unit and a.unit != b.unit):
            mismatch += 1

    missing_any = sum(
        1 for r in records_sorted if not (r.purpose and r.measure_range and r.unit)
    )
    print(f"pdf_used={pdf_path2}")
    if xlsx_path2 and xlsx_path2.exists():
        print(f"xlsx_used={xlsx_path2}")
    print(f"pdf_rows={len(pdf_map)}")
    print(f"xlsx_rows={len(xlsx_map)}")
    print(f"union_rows={len(records_sorted)}")
    print(f"overlap_rows={len(overlap)}")
    print(f"overlap_with_any_nonempty_field_mismatch={mismatch}")
    print(f"rows_with_missing_any_field_after_merge={missing_any}")


if __name__ == "__main__":
    main()
