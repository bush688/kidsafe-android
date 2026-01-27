from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl
from openpyxl.styles import PatternFill


def _norm_tag(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s+", "", s)
    return s


_TAG_EQUIV_CODE_CANON: dict[str, str] = {
    "PI": "PT",
    "PT": "PT",
    "TI": "TE",
    "TE": "TE",
    "LI": "LT",
    "LT": "LT",
    "VI": "VE",
    "VE": "VE",
}


def _canon_tag(v: object) -> str:
    s = _norm_tag(v)
    if s == "":
        return ""

    m = re.match(r"^([A-Z]*\d+)([A-Z]{2})([_-].*)$", s.upper())
    if not m:
        return s.upper()

    prefix, code, rest = m.group(1), m.group(2), m.group(3)
    canon_code = _TAG_EQUIV_CODE_CANON.get(code, code)
    return f"{prefix}{canon_code}{rest}"


def _to_decimal(v: object) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = str(v).strip()
    if s == "":
        return None
    s2 = s.replace(",", "")
    try:
        return Decimal(s2)
    except InvalidOperation:
        return None


def _format_pair(low: object, high: object) -> str:
    def fmt(x: object) -> str:
        if x is None:
            return ""
        if isinstance(x, float):
            return f"{x:.10g}"
        return str(x).strip()

    return f"下限={fmt(low)}, 上限={fmt(high)}"


@dataclass(frozen=True)
class HeaderInfo:
    header_row: int
    col_tag: int
    col_low: int
    col_high: int


def _find_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> HeaderInfo:
    tag_keys = {"位号", "tag", "t ag"}
    low_keys = {"量程下限", "下限", "lrv", "range low", "low range"}
    high_keys = {"量程上限", "上限", "urv", "range high", "high range"}

    def norm_header(x: object) -> str:
        if x is None:
            return ""
        s = str(x).strip().lower()
        s = re.sub(r"\s+", "", s)
        return s

    for r in range(1, 21):
        headers = [norm_header(ws.cell(row=r, column=c).value) for c in range(1, 101)]
        if all(h == "" for h in headers[:8]):
            continue

        def find_col(keys: set[str]) -> Optional[int]:
            for idx, h in enumerate(headers, start=1):
                if not h:
                    continue
                for k in keys:
                    if k in h:
                        return idx
            return None

        col_tag = find_col(tag_keys)
        col_low = find_col(low_keys)
        col_high = find_col(high_keys)
        if col_tag and col_low and col_high:
            return HeaderInfo(header_row=r, col_tag=col_tag, col_low=col_low, col_high=col_high)

    raise ValueError("未在前20行内找到表头：需要包含“位号/量程下限/量程上限”列。")


def _iter_data_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    col_tag: int,
) -> Iterable[int]:
    max_row = ws.max_row or header_row
    for r in range(header_row + 1, max_row + 1):
        tag = _norm_tag(ws.cell(row=r, column=col_tag).value)
        if tag == "":
            continue
        yield r


def _load_compare_map(
    xlsx_path: Path,
    sheet_name: Optional[str],
) -> tuple[dict[str, tuple[object, object]], HeaderInfo, str]:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        chosen_name = sheet_name or ("数据" if "数据" in wb.sheetnames else wb.sheetnames[0])
        ws = wb[chosen_name]
        headers = _find_headers(ws)

        m: dict[str, tuple[object, object]] = {}
        max_col = max(headers.col_tag, headers.col_low, headers.col_high)
        for row in ws.iter_rows(
            min_row=headers.header_row + 1,
            max_row=ws.max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ):
            if not row:
                continue
            tag_raw = row[headers.col_tag - 1] if headers.col_tag - 1 < len(row) else None
            tag = _canon_tag(tag_raw)
            if tag == "":
                continue
            low = row[headers.col_low - 1] if headers.col_low - 1 < len(row) else None
            high = row[headers.col_high - 1] if headers.col_high - 1 < len(row) else None
            if tag not in m:
                m[tag] = (low, high)
        return m, headers, chosen_name
    finally:
        wb.close()


def compare_and_mark(
    base_path: Path,
    compare_path: Path,
    out_dir: Optional[Path] = None,
    base_sheet: Optional[str] = None,
    compare_sheet: Optional[str] = None,
) -> Path:
    if not base_path.exists():
        raise FileNotFoundError(f"基准文件不存在：{base_path}")
    if not compare_path.exists():
        raise FileNotFoundError(f"比对文件不存在：{compare_path}")

    compare_map, _, chosen_compare_sheet = _load_compare_map(compare_path, compare_sheet)

    wb = openpyxl.load_workbook(str(base_path))
    try:
        chosen_base_sheet = base_sheet or wb.sheetnames[0]
        ws = wb[chosen_base_sheet]
        headers = _find_headers(ws)

        green_fill = PatternFill("solid", fgColor="C6EFCE")
        red_fill = PatternFill("solid", fgColor="FFC7CE")

        col_base_out = 6
        col_cmp_out = 7
        ws.cell(row=headers.header_row, column=col_base_out).value = "基准量程(下限/上限)"
        ws.cell(row=headers.header_row, column=col_cmp_out).value = "比对量程(下限/上限)"

        header_style_src = ws.cell(row=headers.header_row, column=max(1, headers.col_high))._style
        ws.cell(row=headers.header_row, column=col_base_out)._style = header_style_src
        ws.cell(row=headers.header_row, column=col_cmp_out)._style = header_style_src

        mismatches = 0
        matched = 0

        for r in _iter_data_rows(ws, headers.header_row, headers.col_tag):
            tag = _canon_tag(ws.cell(row=r, column=headers.col_tag).value)
            if tag == "":
                continue

            cmp_pair = compare_map.get(tag)
            if cmp_pair is None:
                continue

            base_low = ws.cell(row=r, column=headers.col_low).value
            base_high = ws.cell(row=r, column=headers.col_high).value
            cmp_low, cmp_high = cmp_pair

            base_low_d = _to_decimal(base_low)
            base_high_d = _to_decimal(base_high)
            cmp_low_d = _to_decimal(cmp_low)
            cmp_high_d = _to_decimal(cmp_high)

            low_equal = (base_low_d == cmp_low_d) if (base_low_d is not None and cmp_low_d is not None) else (
                str(base_low).strip() == str(cmp_low).strip()
            )
            high_equal = (base_high_d == cmp_high_d) if (base_high_d is not None and cmp_high_d is not None) else (
                str(base_high).strip() == str(cmp_high).strip()
            )

            matched += 1

            if low_equal and high_equal:
                ws.cell(row=r, column=headers.col_low).fill = green_fill
                ws.cell(row=r, column=headers.col_high).fill = green_fill
                ws.cell(row=r, column=col_base_out).value = None
                ws.cell(row=r, column=col_cmp_out).value = None
                continue

            if low_equal:
                ws.cell(row=r, column=headers.col_low).fill = green_fill
            else:
                ws.cell(row=r, column=headers.col_low).fill = red_fill
            if high_equal:
                ws.cell(row=r, column=headers.col_high).fill = green_fill
            else:
                ws.cell(row=r, column=headers.col_high).fill = red_fill

            ws.cell(row=r, column=col_base_out).value = _format_pair(base_low, base_high)
            ws.cell(row=r, column=col_cmp_out).value = _format_pair(cmp_low, cmp_high)
            mismatches += 1

        out_dir2 = out_dir or base_path.parent
        out_dir2.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir2 / f"{base_path.stem}_比对结果_{ts}{base_path.suffix}"
        wb.save(str(out_path))

        print(
            f"完成：基准Sheet={chosen_base_sheet}；比对Sheet={chosen_compare_sheet}；"
            f"位号匹配={matched}；不一致={mismatches}；输出={out_path}"
        )
        return out_path
    finally:
        wb.close()


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description="按位号(Tag)比对两份Excel量程上下限，并对基准文件着色输出。")
    p.add_argument("--base", required=True, help="基准文件路径（xlsx）")
    p.add_argument("--compare", required=True, help="比对文件路径（xlsx）")
    p.add_argument("--out-dir", default="", help="输出目录（默认：基准文件所在目录）")
    p.add_argument("--base-sheet", default="", help="基准Sheet名（默认：第一个Sheet）")
    p.add_argument("--compare-sheet", default="", help="比对Sheet名（默认：优先“数据”，否则第一个Sheet）")
    args = p.parse_args(argv)

    out_dir = Path(args.out_dir) if args.out_dir else None
    base_sheet = args.base_sheet or None
    compare_sheet = args.compare_sheet or None

    try:
        compare_and_mark(
            base_path=Path(args.base),
            compare_path=Path(args.compare),
            out_dir=out_dir,
            base_sheet=base_sheet,
            compare_sheet=compare_sheet,
        )
        return 0
    except Exception as e:
        print(f"失败：{type(e).__name__}: {e}")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
