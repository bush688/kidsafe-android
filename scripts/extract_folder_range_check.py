from __future__ import annotations

import argparse
import asyncio
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Font
from winrt.windows.graphics.imaging import BitmapDecoder
from winrt.windows.media.ocr import OcrEngine
from winrt.windows.storage.streams import InMemoryRandomAccessStream, DataWriter


def _norm(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s))


def _clean_ocr_text(s: str) -> str:
    s2 = re.sub(r"[ \t\r\n]+", "", s or "")
    s2 = s2.replace("，", ",").replace("。", ".")
    return s2.strip()


@dataclass(frozen=True)
class Record:
    instrument_tag: str
    purpose: str
    measure_range: str
    unit: str
    source_file: str


TARGET_HEADERS = ["仪表位号", "用途", "测量范围", "工程单位"]


def _write_excel(out_path: Path, records: Iterable[Record]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "提取结果"

    ws.append(["仪表位号", "用途", "测量范围", "工程单位"])
    for r in records:
        ws.append([r.instrument_tag, r.purpose, r.measure_range, r.unit])

    header_font = Font(bold=True)
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    widths = [18, 28, 26, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


async def _ocr_png_bytes(png_bytes: bytes) -> list[tuple[str, float, float, float, float]]:
    stream = InMemoryRandomAccessStream()
    writer = DataWriter(stream)
    writer.write_bytes(png_bytes)
    await writer.store_async()
    stream.seek(0)

    decoder = await BitmapDecoder.create_async(stream)
    bmp = await decoder.get_software_bitmap_async()

    engine = OcrEngine.try_create_from_user_profile_languages()
    if engine is None:
        return []
    res = await engine.recognize_async(bmp)

    words: list[tuple[str, float, float, float, float]] = []
    for line in res.lines:
        for w in line.words:
            r = w.bounding_rect
            words.append((w.text or "", r.x, r.y, r.width, r.height))
    return words


def _words_in_bbox(
    words: list[tuple[str, float, float, float, float]],
    bbox_px: tuple[float, float, float, float],
) -> str:
    x0, top, x1, bottom = bbox_px
    chosen: list[tuple[float, str]] = []
    for text, x, y, w, h in words:
        cx = x + w / 2
        cy = y + h / 2
        if x0 <= cx <= x1 and top <= cy <= bottom:
            chosen.append((x, text))
    chosen.sort(key=lambda t: t[0])
    return "".join(t for _, t in chosen).strip()


def _pick_main_table(page: pdfplumber.page.Page):
    tables = page.find_tables(
        {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "intersection_tolerance": 5,
            "snap_tolerance": 3,
            "join_tolerance": 3,
            "edge_min_length": 10,
            "min_words_vertical": 1,
            "min_words_horizontal": 1,
        }
    )
    if not tables:
        return None
    tables_sorted = sorted(tables, key=lambda t: len(t.cells), reverse=True)
    return tables_sorted[0]


def _extract_from_pdf_scanned(
    pdf_path: Path,
    resolution: int = 180,
) -> tuple[list[Record], dict[str, int]]:
    records: dict[str, Record] = {}
    stats = {
        "pages_total": 0,
        "pages_with_table": 0,
        "pages_with_header": 0,
        "rows_emitted": 0,
        "rows_skipped_no_tag": 0,
    }

    header_cols: Optional[dict[str, int]] = None
    header_row_top_px: Optional[float] = None

    tag_re = re.compile(r"^[0-9A-Z]{2,}[-0-9A-Z]{3,}$")

    with pdfplumber.open(str(pdf_path)) as pdf:
        stats["pages_total"] = len(pdf.pages)
        scale = resolution / 72.0

        for page in pdf.pages:
            table = _pick_main_table(page)
            if table is None:
                continue

            rows = table.rows
            cols = table.columns
            if len(cols) < 4 or len(rows) < 2:
                continue

            stats["pages_with_table"] += 1

            img = page.to_image(resolution=resolution).original
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            words = asyncio.run(_ocr_png_bytes(buf.getvalue()))
            if not words:
                continue

            def cell_bbox_px(r_idx: int, c_idx: int) -> tuple[float, float, float, float]:
                row_bbox = rows[r_idx].bbox
                col_bbox = cols[c_idx].bbox
                x0 = col_bbox[0]
                x1 = col_bbox[2]
                top = row_bbox[1]
                bottom = row_bbox[3]
                return (x0 * scale, top * scale, x1 * scale, bottom * scale)

            if header_cols is None:
                for r_idx in range(min(6, len(rows))):
                    cell_texts = []
                    for c_idx in range(min(len(cols), 12)):
                        t = _words_in_bbox(words, cell_bbox_px(r_idx, c_idx))
                        cell_texts.append(_clean_ocr_text(t))
                    joined = "".join(cell_texts)
                    if all(h in joined for h in TARGET_HEADERS):
                        header_cols = {}
                        for c_idx, t in enumerate(cell_texts):
                            for h in TARGET_HEADERS:
                                if h in t and h not in header_cols:
                                    header_cols[h] = c_idx
                        if len(header_cols) == 4:
                            header_row_top_px = rows[r_idx].bbox[1] * scale
                            stats["pages_with_header"] += 1
                            break

            if header_cols is None:
                continue

            start_row = 0
            if header_row_top_px is not None:
                for i, r in enumerate(rows):
                    if r.bbox[1] * scale >= header_row_top_px - 1:
                        start_row = i + 1
                        break

            for r_idx in range(start_row, len(rows)):
                instrument = _clean_ocr_text(
                    _words_in_bbox(words, cell_bbox_px(r_idx, header_cols["仪表位号"]))
                )
                instrument = instrument.replace("\\", "-").replace("—", "-").replace("–", "-")
                instrument_n = _norm(instrument)
                if not instrument_n:
                    stats["rows_skipped_no_tag"] += 1
                    continue

                if not tag_re.match(instrument_n):
                    stats["rows_skipped_no_tag"] += 1
                    continue

                purpose = _clean_ocr_text(
                    _words_in_bbox(words, cell_bbox_px(r_idx, header_cols["用途"]))
                )
                measure_range = _clean_ocr_text(
                    _words_in_bbox(words, cell_bbox_px(r_idx, header_cols["测量范围"]))
                )
                unit = _clean_ocr_text(
                    _words_in_bbox(words, cell_bbox_px(r_idx, header_cols["工程单位"]))
                )

                if instrument_n not in records:
                    records[instrument_n] = Record(
                        instrument_tag=instrument,
                        purpose=purpose,
                        measure_range=measure_range,
                        unit=unit,
                        source_file=pdf_path.name,
                    )
                    stats["rows_emitted"] += 1

    return list(records.values()), stats


def _discover_pdfs(folder: Path) -> list[Path]:
    pdfs = sorted(folder.glob("*.pdf"))
    prefer: dict[str, Path] = {}
    for p in pdfs:
        k = re.sub(r"\s+", "", p.name)
        k = k.replace("222.pdf", "").replace("333.pdf", "")
        if k not in prefer:
            prefer[k] = p
        else:
            if "333" in p.name and "333" not in prefer[k].name:
                prefer[k] = p
    return sorted(set(prefer.values()), key=lambda x: x.name)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input_dir", type=str, required=True)
    ap.add_argument("--out", type=str, required=True)
    ap.add_argument("--resolution", type=int, default=180)
    args = ap.parse_args()

    folder = Path(args.input_dir)
    if not folder.exists():
        raise FileNotFoundError(f"目录不存在：{folder}")

    pdfs = _discover_pdfs(folder)
    if not pdfs:
        raise FileNotFoundError(f"目录下未找到PDF：{folder}")

    all_records: dict[str, Record] = {}
    totals = {
        "files": 0,
        "pages_total": 0,
        "pages_with_table": 0,
        "pages_with_header": 0,
        "rows_emitted": 0,
        "rows_skipped_no_tag": 0,
    }

    for pdf_path in pdfs:
        try:
            recs, st = _extract_from_pdf_scanned(pdf_path, resolution=args.resolution)
        except Exception:
            continue

        totals["files"] += 1
        for k in totals:
            if k in st:
                totals[k] += int(st[k])

        for r in recs:
            key = _norm(r.instrument_tag)
            if key and key not in all_records:
                all_records[key] = r

    if not all_records:
        raise RuntimeError("未能从目录PDF中提取到有效记录（可能需要安装/启用系统OCR语言包）。")

    out_path = Path(args.out)
    records_sorted = sorted(all_records.values(), key=lambda r: _norm(r.instrument_tag))
    _write_excel(out_path, records_sorted)

    missing_any = sum(
        1
        for r in records_sorted
        if not (r.purpose and r.measure_range and r.unit)
    )
    print(f"files_processed={totals['files']}")
    print(f"pages_total={totals['pages_total']}")
    print(f"pages_with_table={totals['pages_with_table']}")
    print(f"pages_with_header={totals['pages_with_header']}")
    print(f"rows_emitted={totals['rows_emitted']}")
    print(f"unique_instrument_tags={len(records_sorted)}")
    print(f"rows_with_missing_any_field={missing_any}")
    print(f"out={out_path}")


if __name__ == "__main__":
    main()
